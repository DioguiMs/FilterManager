# logic.py

from ast import Str
import os
import sys
from types import new_class
import openpyxl
import re
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtWidgets import QPushButton
from openpyxl.utils import exceptions


#### File dialog

def open_dialog(self, button_name, button, file_filter, rogers_path_label, rq_path_label):
    file_path, _ = QFileDialog.getOpenFileName(button, "Select File", filter=file_filter)
    print("File path:", file_path)
    file_name = os.path.basename(file_path)  
    
    global rogers_file_path 
    global rq_file_path
        
    
    if button_name == "rogers":
        
        rogers_file_path = file_path
        rogers_file_name = file_name
        
        print("Rogers file path:", rogers_file_path)
        print("Rogers file name:", rogers_file_name)
        rogers_path_label.setText("File selected: " + file_name)      

        
    elif button_name == "rq":
        
        rq_file_path = file_path
        rq_file_name = file_name
        
        print("RQ file path:", rq_file_path)
        print("RQ file name:", rq_file_name)
        rq_path_label.setText("File selected: " + file_name)     
    

#### Run button logic
            
class PriceFilter(QPushButton):
    def __init__(self, parent=None):        
        super(PriceFilter, self).__init__(parent)
        
        if rogers_file_path in [None, ''] and rq_file_path in [None, '']:
            print("Both file paths are missing")
            
        elif rogers_file_path in [None, '']:
            print("Rogers file path is missing")
            
        elif rq_file_path in [None, '']:
            print("RQ file path is missing")
        
        else:            
            self.filter_function()
            
                        
            

    # filter_function --brain--

    def filter_function(self):
        
        files = False
        
        rogers_workbook = None
        rq_workbook = None
        
        try:            
            rogers_workbook = openpyxl.load_workbook(rogers_file_path)
            rq_workbook = openpyxl.load_workbook(rq_file_path)
            
            files = True
            
        except FileNotFoundError:
            print("File not found. Please check the file path.")
        
        except exceptions.InvalidFileException:
            print("Invalid file format. Please make sure the file is in a supported format (e.g., .xlsx, .xlsm).")
            
        except Exception as e:
            print("An error occurred while loading the workbook:", str(e))
            
        if files is True and rogers_workbook is not None and rq_workbook is not None:
            
            pattern = r'^SKU$'
            ignore_pattern_case = re.compile(pattern, re.IGNORECASE)
            
            global rogers_sheet
            global rq_sheet
            
            rogers_sheet = rogers_workbook['Consumer Hardware Pricing']
            rq_sheet = rq_workbook["Hardware"]
            
            sku_cords = None
            for row in rogers_sheet.iter_rows():                
                for cell in row:
                    if cell.value is not None and ignore_pattern_case.search(str(cell.value)):
                        print(cell.value)
                        print("row:", cell.row)
                        print("column:", cell.column)
                        sku_cords = [cell.row, cell.column]
                        break
                    
                if sku_cords is not None: break
                            
                    
            if not sku_cords:
                print("Error: SKU cell was not found")
                
            if sku_cords:
                
                sku_row = sku_cords[0]
                sku_column = sku_cords[1]
                rq_column = None
                headers = None
                rebate_headers = None
                ranges = None
                RPP_headers = None
                watch_headers = None
                tablet_headers = None
                
                pattern = r'(.*)(NRS$)'
                ignore_pattern_case = re.compile(pattern, re.IGNORECASE)
                
                
                for row in rogers_sheet.iter_rows(min_row=sku_row+1, min_col=sku_column ,max_col=sku_column):
                    for cell in row:
                        sku_number = cell.value
                        device_cords = [cell.row, cell.column]    
                        
                        # First run to encounter the correct column
                        if rq_column is None and sku_number is not None:
                            for row in rq_sheet.iter_rows():
                                for cell in row:
                                    
                                    rq_sku = cell.value
                                    
                                    if str(rq_sku).lower() == str(sku_number).lower():
                                        print("SKU was found in the RQ sheet, copying it's values....")
                                        
                                        rq_sku_cords = [cell.row, cell.column]

                                        if headers is None:
                                            headers, rebate_headers, ranges, RPP_headers, watch_headers, tablet_headers = headers_list()
                                        copy(device_cords, rq_sku_cords, headers, rebate_headers, ranges, RPP_headers, watch_headers, tablet_headers, sku_number)
                                        
                                        if len(headers) < 4:
                                            print("Error the amount of pattern founds were less than expected, The values found were:")
                                            value_number = 1
                                            for header in headers:
                                                print(f"value {value_number}: {header}")
                                                value_number += 1
                                        
                                        rq_column = rq_sku_cords[1]
                                        if rq_column is not None: break
                                        
                                if rq_column is not None: break
                        
                        # Rest of the runs once the column was found        
                        if rq_column is not None and sku_number is not None:
                            for row in rq_sheet.iter_rows(min_col=rq_column, max_col=rq_column):
                                for cell in row:
                                    
                                    rq_sku = cell.value
                                    
                                    if ignore_pattern_case.match(str(rq_sku)):
                                        matches = re.findall(pattern, str(rq_sku), re.IGNORECASE)
                                        
                                        rq_sku = matches[0][0] # type: ignore
                                    
                                    
                                    if str(rq_sku).lower() == str(sku_number).lower():
                                        print("SKU was found in the RQ sheet, copying it's values....")
                                        rq_sku_cords = [cell.row, cell.column]

                                        if headers is None:
                                            headers, rebate_headers, ranges, RPP_headers, watch_headers, tablet_headers = headers_list()
                                        copy(device_cords, rq_sku_cords, headers, rebate_headers, ranges, RPP_headers, watch_headers, tablet_headers, sku_number)
                
                
                ############## Save File dialog
                
                save_dialog = QFileDialog()
                save_dialog.setDefaultSuffix(".xlsx")
                save_dialog.setNameFilters(["Excel files (*.xlsx)"])
                
                file_name = os.path.basename(rq_file_path)
                
                save_dialog.selectFile(file_name)
                
                save_path, _ = save_dialog.getSaveFileName(None, "Save File", "", "Excel files (*.xlsx)")
                
                if save_path:
                    
                    rq_workbook.save(save_path)
                    ("File saved successfully!")
                    
                else:
                    print("File save canceled")


def headers_list():
    
    ### Rogers Pricing sheet headers
    
    patterns = [
        (r'^MSRP$', "MSRP"), # MSRP #1 of the list [0]
        (r'RPP\s*\W\s*FINANCING', "RPP Financing"), # RPP - Financing is #2 of the list [1]
        (r'RPP(?:.*)Talk(?:.*)Text(?:.*)Financing', "RPP Talk and Text Financing"), # RPP Talk and Text is #3 of the list [2]
        (r'RPP(?:.*)DATA(?:.*)ONLY(?:.*)FINANCING', "RPP Data Only Financing") # RPP Data Only Financing is #4 of the list [3]
    ]
    
    headers = []
    
    header_row = None
    for pattern, label in patterns:
        ignore_pattern_case = re.compile(pattern, re.IGNORECASE)
        
        stop = False
        for row in rogers_sheet.iter_rows(min_row=header_row, max_col=header_row):
            for cell in row:
                if cell.value is not None and ignore_pattern_case.search((str(cell.value))):
                    stop = True
                    
                    headers.append(cell)
                    
                    if stop == True: break
                
            if stop == True: break

                    
    print(f"MSRP coords are row: {headers[0].row}, column: {headers[0].column_letter}")
    print(f"RPP Financing coords are row: {headers[1].row}, column: {headers[1].column_letter}")
    print(f"RPP Talk and Text Financing coords are row: {headers[2].row}, column: {headers[2].column_letter}")
    print(f"RPP Data Only Financing coords are row: {headers[3].row}, column: {headers[3].column_letter}")
    
    
    # Rogers Pricing sheet Ranges
    
    patterns = [
        (r'^upfront.*edge$'),        # Upfront edge pattern
        (r'^Discount.*Duration$')   # Discount duration pattern
    ]
    
    device_pricing_range = []
    talktext_pricing_range = []
    tablet_pricing_range = []
    
    
    UpfrontEdgeSearch = re.compile(patterns[0], re.IGNORECASE)
    DiscountDuration = re.compile(patterns[1], re.IGNORECASE)
    
    
    total = 0
    for row in rogers_sheet.iter_rows(min_col=headers[1].column, min_row=headers[1].row):
        for cell in row:
            if cell.value is not None and UpfrontEdgeSearch.search((str(cell.value))):
                
                device_pricing_range.append(cell)
                total += 1
                if total == 2: break
            
        if total == 2: break
    
    stop = False
    for row in rogers_sheet.iter_rows(min_col=headers[2].column, min_row=headers[2].row):
        for cell in row:
            if cell.value is not None and DiscountDuration.search((str(cell.value))):
                
                talktext_pricing_range.append(cell)
                stop = True
                break
            
        if stop == True: break
    
    stop= False
    for row in rogers_sheet.iter_rows(min_col=headers[3].column, min_row=headers[3].row):
        for cell in row:
            if cell.value is not None and DiscountDuration.search((str(cell.value))):
                
                tablet_pricing_range.append(cell)
                stop = True
                break
            
        if stop == True: break
        
    ranges = [
        device_pricing_range,
        talktext_pricing_range,
        tablet_pricing_range        
    ]
    
    
    ################# RQ Headers

    
    patterns = [
        (r'RPP(?!.*VOICE).*DATA.*'), # Pattern for tablets (matches with rebate) [0]
        (r'RPP.+?watch.+'), # Pattern for watches (matches with rebates) [1]
        (r'RPP.+'), # Takes all the rest [2]
        (r'RPP.+?Rebate.+') # Rebate pattern [3]
    ]
    
    header_row = None
    stop = False

    tablet_headers = []
    watch_headers = []
    RPP_headers = []
    
    tablet_rebate_headers = []
    watch_rebate_headers = []
    RPP_rebate_headers = []
    rebate_headers = [RPP_rebate_headers, tablet_rebate_headers, watch_rebate_headers]
    
    rebate_search = re.compile(patterns[3], re.IGNORECASE)
    tablet_search = re.compile(patterns[0], re.IGNORECASE)
    watch_search = re.compile(patterns[1], re.IGNORECASE)
    rpp_search = re.compile(patterns[2], re.IGNORECASE)
    
    for row in rq_sheet.iter_rows(min_row=header_row, max_row=header_row):
        for cell in row:
            if cell is not None and tablet_search.search((str(cell.value))):
                
                header_row = cell.row
                
                if rebate_search.search((str(cell.value))):

                    rebate_headers[1].append(cell)                    
                    
                else: 
                
                    tablet_headers.append(cell)
                    
            if cell is not None and cell not in tablet_headers and not any(cell in rebates for rebates in rebate_headers) and watch_search.search((str(cell.value))):
                
                header_row = cell.row
                
                if rebate_search.search((str(cell.value))):
                    
                    rebate_headers[2].append(cell)
                    
                else:
                    
                    watch_headers.append(cell)
                    
            if cell is not None and cell not in tablet_headers and cell not in watch_headers and not any(cell in rebates for rebates in rebate_headers) and rpp_search.search((str(cell.value))):
                
                header_row = cell.row
                
                if rebate_search.search((str(cell.value))):
                    
                    rebate_headers[0].append(cell)
                    
                else:
                    
                    RPP_headers.append(cell)

                    
                    
                
                    
    return headers, rebate_headers, ranges, RPP_headers, watch_headers, tablet_headers




def copy(device_cords, rq_sku_cords, headers, rebate_headers, ranges, RPP_headers, watch_headers, tablet_headers, sku_number):
    
    # headers:
    #    (r'^MSRP$', "MSRP"), # MSRP #1 of the list [0]
    #    (r'RPP\s*\W\s*FINANCING', "RPP Financing"), # RPP - Financing is #2 of the list [1]
    #    (r'RPP(?:.*)Talk(?:.*)Text(?:.*)Financing', "RPP Talk and Text Financing"), # RPP Talk and Text is #3 of the list [2]
    #    (r'RPP(?:.*)DATA(?:.*)ONLY(?:.*)FINANCING', "RPP Data Only Financing") # RPP Data Only Financing is #4 of the list [3]
    
    # rebate_headers = [RPP_rebate_headers, tablet_rebate_headers, watch_rebate_headers]
    
    # ranges = [device_pricing_range, talktext_pricing_range, tablet_pricing_range]
    
    pattern = r'^aw.+'
    
    pattern_ignore_case = re.compile(pattern, re.IGNORECASE)
    
    watch = False
    if pattern_ignore_case.search((str(sku_number))):
        watch = True
    
    
    
    msrp  = rogers_sheet.cell(row=device_cords[0], column=headers[0].column).value

    def range_check(row_range, column_range):
        
        value = 0        
        for row in rogers_sheet.iter_rows(min_row=row_range[0], max_row=row_range[1], min_col=column_range[0], max_col=column_range[1]):
            for cell in row:
                if cell.value not in [None, '', '-']:
                    print('cell has a value: ', cell.value)
                    value += 1

                elif cell.value == '-':
                    print('cell has no value')

                else:
                    print('unknown')
                    
        return value
    
    
    def transpose(value, main, rebate, msrp):
        
        if value > 0:
            for cell in main:
                edited_cell = rq_sheet.cell(row=rq_sku_cords[0], column=cell.column)
                edited_cell.value = float(msrp)
                #rq_sheet.cell(row=rq_sku_cords[0], column=cell.column).value = float(msrp)

            for cell in rebate:
                edited_cell = rq_sheet.cell(row=rq_sku_cords[0], column=cell.column)
                edited_cell.value = int(0)
                #rq_sheet.cell(row=rq_sku_cords[0], column=cell.column).value = int(0)

        else:
            for cell in main:
                edited_cell = rq_sheet.cell(row=rq_sku_cords[0], column=cell.column)
                edited_cell.value = str("N")
                #rq_sheet.cell(row=rq_sku_cords[0], column=cell.column).value = str('N')

            for cell in rebate:
                edited_cell = rq_sheet.cell(row=rq_sku_cords[0], column=cell.column)
                edited_cell.value = str("N")
                #rq_sheet.cell(row=rq_sku_cords[0], column=cell.column).value = str('N')
    
    #### National RPP financing
    
    row_range = [device_cords[0], device_cords[0]]
    column_range = [headers[1].column, ranges[0][0].column]
    
    value = range_check(row_range, column_range)
    transpose(value, RPP_headers, rebate_headers[0], msrp)
    value = 0
    
    ##### Data Only discovery
    
    column_range = [headers[3].column, ranges[2][0].column]
    value = range_check(row_range, column_range)
    
    
    #### Watch financing
    
    if watch is True:
        transpose(value, watch_headers, rebate_headers[2], msrp)
        transpose(0, tablet_headers, rebate_headers[1], msrp)
        
    if watch is False:
        transpose(value, tablet_headers, rebate_headers[1], msrp)
        transpose(0, watch_headers, rebate_headers[2], msrp)
